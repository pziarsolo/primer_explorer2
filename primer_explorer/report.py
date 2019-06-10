from primer_explorer.config import SHORT_PRODUCTS_CUTOFF

from openpyxl import Workbook
from _collections import OrderedDict
from primer_explorer.stats import (NUM_SEQUENCIABLE_PRODUCTS,
                                   NUM_OF_POSSIBLE_PRODUCTS_700,
                                   NUM_OF_POSSIBLE_PRODUCTS_10000,
                                   NUM_SHORT_PRODUCTS,
                                   NUM_UNIQUE_REPETITIVE_PRODUCTS,
                                   NUM_UNIQUE_UNIQUE_PRODUCTS,
                                   NUM_REPETITIVE_REPETITIVE_PRODUCTS,
                                   ADJUSTED_PERCENTAGE_OF_SEQUENCIABLE_NUCLEOTIDES,
                                   PERCENTAGE_OF_SEQUENCIABLE_NUCLEOTIDES,
                                   NUM_UNION_SITES_P1, NUM_UNION_SITES_P2)

from openpyxl.styles import Font, colors
from openpyxl.utils.cell import get_column_letter

LABELS_TO_REPORT = OrderedDict([
    (NUM_UNION_SITES_P1, {'code': 'P1_COUNT',
                               'label': 'Num Union sites os primer1'}),
    (NUM_UNION_SITES_P2, {'code': 'P2_COUNT',
                          'label': 'Num Union sites os primer2'}),
    (NUM_OF_POSSIBLE_PRODUCTS_10000, {'code': 'NP10K',
                                      'label': 'Num of possible products(length < 10000'}),
    (NUM_OF_POSSIBLE_PRODUCTS_700, {'code': 'NP700',
                                    'label': 'Num of possible products(length < 700'}),
    (NUM_SHORT_PRODUCTS, {'code': 'NSP',
                          'label': 'Num of short products(length < 100'}),
    (NUM_SEQUENCIABLE_PRODUCTS, {'code': 'NSECP',
                                 'label': 'Num of sequentiable products(100-700)'}),
    (NUM_UNIQUE_UNIQUE_PRODUCTS, {'code': 'NUUP',
                                  'label': 'Num of unique unique products (100-700)'}),
    (NUM_REPETITIVE_REPETITIVE_PRODUCTS, {'code': 'NUUP',
                                          'label': 'Num of repetitive repetitive products (100-700)'}),
    (NUM_UNIQUE_REPETITIVE_PRODUCTS, {'code': 'NURP',
                                      'label': 'Num of unique repetitive products (100-700)'}),
    (PERCENTAGE_OF_SEQUENCIABLE_NUCLEOTIDES, {'code': 'PSN',
                                              'label': 'Percentage of sequenciable nucleotides'}),
    (ADJUSTED_PERCENTAGE_OF_SEQUENCIABLE_NUCLEOTIDES, {'code': 'APSN',
                                                       'label': 'Adjusted Percentage of sequenciable nucleotides'})
])


def generate_pair_stats(primer_pair, count, short_products_cutoff=SHORT_PRODUCTS_CUTOFF):
    report = []
    report.append("{},{}".format(primer_pair[0].decode(), primer_pair[1].decode()))
    report.append("-" * 20)
    report.append("Total Number of products:\t{}".format(count["total_products"]))

    amplificable_product_count = count['amplificable_products']
    report.append("Pcr products that are amplificable(<1000): {}".format(amplificable_product_count))
    viable_products_count = count["viable_products"]['count']
    viable_products_ratio = float(viable_products_count / amplificable_product_count)

    text = "Number of pcr products in range ({} - {}):\t{}\t{:.1%}"
    text = text.format(count["viable_products"]['min'],
                       count["viable_products"]['max'],
                       viable_products_count, viable_products_ratio)
    report.append(text)

    short_product_counts = count['filtered_by_max_length']['count']
    text = "Number of short products(<{}):\t{}\t{:.1%}"
    text = text.format(count['filtered_by_max_length']['max'],
                       short_product_counts,
                       short_product_counts / amplificable_product_count)
    report.append(text)

    report.append("Number of euchromatic effective products:\t{}".format(count["euchromatin_products"]))
    report.append("Number of effective euchromatic nucleotides:\t{}".format(count["euchromatin_nucleotides"]))
    report.append("Number of heterochromatic effective products:\t{}".format(count["heterochromatin_products"]))
    report.append("Number of effective heterochromatic nucleotides:\t{}".format(count["heterochromatin_nucleotides"]))
    report.append("Number of mixed effective products:\t{}".format(count["mixed_products"]))
    report.append("Number of mixed effective nucleotides:\t{}".format(count["mixed_nucleotides"]))
    report.append("Ratio euchromatin (without mixed):\t{:.1%}".format(count["euchromatin_products"] / (viable_products_count - count["mixed_products"])))
    report.append("Ratio heterochromatin (without mixed):\t{:.1%}".format(count["heterochromatin_products"] / (viable_products_count - count["mixed_products"])))
    if viable_products_ratio <= short_products_cutoff:
        report.append("WARNING: too much products generated by this pair are no effective")
    return report


def write_detailed_report(report_fhand, stats):
    report = []
    for idx, set_stats in stats.items():
        primers = set_stats['primers']
        report.append("PRIMER SET {}".format(str(idx)))
        report.append('Primers: ' + ', '.join([p.decode() for p in primers]))
        report.append("#" * 30)
        for pair in sorted(set_stats['stats'].keys()):
            counts = set_stats['stats'][pair]
            report.extend(generate_pair_stats(pair, counts))
            report.append("-" * 20)

    report_fhand.write("\n".join(report))
    report_fhand.flush()


def write_gff_report(gff_results, output_fhand):
    report = []
    report.append("PRIMER PAIR\tEXONS\tGENES\tNUM_PCR_PRODUCTS")
    for primer_pair, results in gff_results.items():
        report.append("{}\t{}\t{}\t{}".format(primer_pair, results["exon"],
                                              results["gene"],
                                              results["num_pcr_products"]))

    output_fhand.write("\n".join(report).encode())
    output_fhand.flush()
    output_fhand.close()


def write_stats_in_excel(out_fpath, stats):
    workbook = Workbook()
    for set_index, primer_set_stats in stats.items():
        if set_index == 0:
            sheet = workbook.active
        else:
            sheet = workbook.create_sheet(title='sheet {}'.format(set_index + 1))
        write_set_stats_in_sheet(primer_set_stats, sheet)
        workbook.save(out_fpath)
        break


def write_set_stats_in_sheet(primer_set_stats, sheet):
    stats = primer_set_stats['stats']
    first = list(stats.values())[0]
    min_sequenciable = first[NUM_SEQUENCIABLE_PRODUCTS]['min']
    max_sequenciable = first[NUM_SEQUENCIABLE_PRODUCTS]['max']
    primers = primer_set_stats['primers']
    labels = OrderedDict([(index + 1, primer) for index, primer in enumerate(primers)])
    table_label = 'Number of sequenciable products between {}-{} bp'
    sheet['A1'] = table_label.format(min_sequenciable, max_sequenciable)
    sheet['G1'] = 'Cells in red: Short products are more than 10% of amplificable'
    sheet['G1'].font = Font(color=colors.RED)

    for index, (primer_index, primer) in enumerate(labels.items()):
        col_index = index + 3
#         print(col_index)
#         print(primer_index)
#         print(primer)
        cell = sheet.cell(column=col_index, row=2, value='Primer_{}'.format(primer_index))
        font = Font(color=colors.BLACK, bold=True)
        cell.font = font
        sheet.cell(column=col_index, row=3, value=primer)

        row_index = index + 4
        cell = sheet.cell(column=1, row=row_index, value='Primer_{}'.format(primer_index))
        font = Font(color=colors.BLACK, bold=True)
        cell.font = font
        sheet.cell(column=2, row=row_index, value=primer)

    already_done = []
    used_combinations = {}
    for index1, primer1 in labels.items():
        for index2, primer2 in labels.items():
            if index1 == index2:
                continue
            if (index1, index2) in already_done:
                continue
            if (index2, index1) in already_done:
                continue
            already_done.append((index1, index2))
            already_done.append((index2, index1))

            counts = stats.get((primer1, primer2), None)
            if counts is None:
                counts = stats.get((primer2, primer1), None)
                used_combinations[(primer2, primer1)] = (index2, index1)
            else:
                used_combinations[(primer1, primer2)] = (index1, index2)

            num_sequenciable_products = counts[NUM_SEQUENCIABLE_PRODUCTS]['count']

            cell = sheet.cell(row=index1 + 3, column=index2 + 2, value=num_sequenciable_products)
            ratio_secuenciable_products = num_sequenciable_products / counts[NUM_OF_POSSIBLE_PRODUCTS_700]
            if ratio_secuenciable_products < 0.90:
                cell.font = Font(color=colors.RED)

    write_legend(sheet)

    write_detailed_stats(sheet, stats, used_combinations)

    for column_index in range(12):
        sheet.column_dimensions[get_column_letter(column_index + 1)].width = 11

    sheet.freeze_panes = sheet['A17']


def write_detailed_stats(sheet, stats, used_combinations, detail_row_index_start=16,
                         detail_column_index_start=1):
    for index, labels in enumerate(LABELS_TO_REPORT.values()):
        code = labels['code']
        sheet.cell(row=detail_row_index_start, column=index + 2, value=code)

    for index, (primers, indexes) in enumerate(used_combinations.items()):
        pair_stats = stats[primers]
        pair_label = 'Pair P{}-P{}'.format(*indexes)
        row_index = index + detail_row_index_start + 1
        sheet.cell(row=row_index, column=detail_column_index_start, value=pair_label)
        for label_index, labels in enumerate(LABELS_TO_REPORT.keys()):
            count = pair_stats[labels]
            try:
                count = count['count']
            except TypeError:
                pass
            value_column_index = detail_column_index_start + label_index + 1
            sheet.cell(row=row_index, column=value_column_index, value=count)

#         print(primers, indexes)


def write_legend(sheet, code_column_index=15, label_column_index=16):
    sheet.cell(row=2, column=code_column_index, value='LEGEND')
    sheet.cell(row=3, column=code_column_index, value='Code')
    sheet.cell(row=3, column=label_column_index, value='Description')

    for index, label in enumerate(LABELS_TO_REPORT.values()):
        sheet.cell(row=index + 4, column=code_column_index, value=label['code'])
        sheet.cell(row=index + 4, column=label_column_index, value=label['label'])
    sheet.column_dimensions[get_column_letter(code_column_index)].width = 11
    sheet.column_dimensions[get_column_letter(label_column_index)].width = 40
